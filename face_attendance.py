import cv2
import face_recognition
import numpy as np
from datetime import datetime
import openpyxl
import os

# Load known images and encode them
def load_images_from_folder(folder):
    images = []
    names = []
    for filename in os.listdir(folder):
        img = cv2.imread(os.path.join(folder, filename))
        if img is not None:
            images.append(img)
            names.append(os.path.splitext(filename)[0])
    return images, names

def encode_images(images):
    encoded_list = []
    for img in images:
        img_rgb = cv2.cvtColor(img, cv2.COLOR_BGR2RGB)
        encode = face_recognition.face_encodings(img_rgb)[0]
        encoded_list.append(encode)
    return encoded_list

# Mark attendance in Excel
def mark_attendance(name):
    wb = openpyxl.load_workbook('Attendance.xlsx')
    sheet = wb.active
    now = datetime.now()
    dt_string = now.strftime('%Y-%m-%d %H:%M:%S')
    sheet.append([name, dt_string])
    wb.save('Attendance.xlsx')

# Load images and encode them
path = 'ImagesAttendance'
images, classNames = load_images_from_folder(path)
encodeListKnown = encode_images(images)

# Initialize webcam
cap = cv2.VideoCapture(0)

while True:
    success, img = cap.read()
    imgS = cv2.resize(img, (0, 0), None, 0.25, 0.25)
    imgS = cv2.cvtColor(imgS, cv2.COLOR_BGR2RGB)

    facesCurFrame = face_recognition.face_locations(imgS)
    encodesCurFrame = face_recognition.face_encodings(imgS, facesCurFrame)

    for encodeFace, faceLoc in zip(encodesCurFrame, facesCurFrame):
        matches = face_recognition.compare_faces(encodeListKnown, encodeFace)
        faceDis = face_recognition.face_distance(encodeListKnown, encodeFace)
        matchIndex = np.argmin(faceDis)

        if matches[matchIndex]:
            name = classNames[matchIndex].upper()
            y1, x2, y2, x1 = faceLoc
            y1, x2, y2, x1 = y1*4, x2*4, y2*4, x1*4
            cv2.rectangle(img, (x1, y1), (x2, y2), (0, 255, 0), 2)
            cv2.putText(img, name, (x1+6, y2-6), cv2.FONT_HERSHEY_COMPLEX, 1, (255, 255, 255), 2)
            mark_attendance(name)
    
    cv2.imshow('Webcam', img)
    if cv2.waitKey(1) & 0xFF == ord('q'):
        break

cap.release()
cv2.destroyAllWindows()

from openpyxl.utils.exceptions import InvalidFileException

def mark_attendance(name):
    try:
        # Attempt to load the workbook
        wb = openpyxl.load_workbook('Attendance.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        # If file does not exist, create a new workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.append(["Name", "Date Time"])  # Adding headers
    except InvalidFileException:
        print("The file is not a valid Excel file.")
        return
    except Exception as e:
        print(f"An error occurred: {e}")
        return

    # Add attendance record
    now = datetime.now()
    dt_string = now.strftime('%Y-%m-%d %H:%M:%S')
    sheet.append([name, dt_string])
    wb.save('Attendance.xlsx')

